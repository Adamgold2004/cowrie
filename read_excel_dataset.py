#!/usr/bin/env python3
"""
Excel Dataset Reader for Cowrie Training

This script attempts to read the Excel dataset using different methods
and extract training patterns for the Cowrie honeypot.
"""

import os
import sys
import json
from datetime import datetime

def try_openpyxl_read(file_path):
    """Try to read Excel file using openpyxl"""
    try:
        from openpyxl import load_workbook
        print("📖 Reading Excel file with openpyxl...")
        
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value))
        
        print(f"📋 Found {len(headers)} columns:")
        for i, header in enumerate(headers):
            print(f"   {i+1:2d}. {header}")
        
        # Read first few rows of data
        data_rows = []
        row_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row_count >= 10:  # Limit to first 10 rows for analysis
                break
            if any(cell is not None for cell in row):
                data_rows.append(row)
                row_count += 1
        
        print(f"\n📊 Sample data ({len(data_rows)} rows):")
        for i, row in enumerate(data_rows[:5]):
            print(f"   Row {i+1}: {row}")
        
        # Extract patterns
        patterns = extract_patterns_from_excel(headers, data_rows, sheet)
        
        workbook.close()
        return patterns
        
    except ImportError:
        print("❌ openpyxl not available")
        return None
    except Exception as e:
        print(f"❌ Error reading with openpyxl: {e}")
        return None

def try_xlrd_read(file_path):
    """Try to read Excel file using xlrd (for older Excel files)"""
    try:
        import xlrd
        print("📖 Reading Excel file with xlrd...")
        
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        
        # Get headers
        headers = []
        if sheet.nrows > 0:
            for col in range(sheet.ncols):
                cell_value = sheet.cell_value(0, col)
                if cell_value:
                    headers.append(str(cell_value))
        
        print(f"📋 Found {len(headers)} columns:")
        for i, header in enumerate(headers):
            print(f"   {i+1:2d}. {header}")
        
        # Read sample data
        data_rows = []
        for row in range(1, min(11, sheet.nrows)):  # First 10 data rows
            row_data = []
            for col in range(sheet.ncols):
                row_data.append(sheet.cell_value(row, col))
            data_rows.append(tuple(row_data))
        
        print(f"\n📊 Sample data ({len(data_rows)} rows):")
        for i, row in enumerate(data_rows[:5]):
            print(f"   Row {i+1}: {row}")
        
        # Extract patterns
        patterns = extract_patterns_from_excel(headers, data_rows, sheet)
        
        return patterns
        
    except ImportError:
        print("❌ xlrd not available")
        return None
    except Exception as e:
        print(f"❌ Error reading with xlrd: {e}")
        return None

def extract_patterns_from_excel(headers, data_rows, sheet_obj):
    """Extract training patterns from Excel data"""
    print("\n🧠 Extracting patterns from Excel data...")

    patterns = {
        'attack_types': set(),
        'network_features': {},
        'port_patterns': set(),
        'traffic_patterns': [],
        'attack_signatures': [],
        'flow_characteristics': {}
    }

    # Find important columns for honeypot training
    attack_type_col = None
    port_col = None
    feature_cols = []

    for i, header in enumerate(headers):
        header_lower = header.lower()
        if 'attack' in header_lower or 'type' in header_lower:
            attack_type_col = i
            print(f"🎯 Found attack type column: {header} (index {i})")
        elif 'port' in header_lower:
            port_col = i
            print(f"🔌 Found port column: {header} (index {i})")
        elif any(keyword in header_lower for keyword in ['packet', 'flow', 'bytes', 'duration', 'flag']):
            feature_cols.append((i, header))

    print(f"📊 Found {len(feature_cols)} network feature columns")

    # Extract attack types and network patterns
    attack_counts = {}
    port_counts = {}

    try:
        row_count = 0
        if hasattr(sheet_obj, 'iter_rows'):  # openpyxl
            for row in sheet_obj.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    row_count += 1

                    # Extract attack type
                    if attack_type_col is not None and len(row) > attack_type_col:
                        attack_type = str(row[attack_type_col]).strip()
                        patterns['attack_types'].add(attack_type)
                        attack_counts[attack_type] = attack_counts.get(attack_type, 0) + 1

                    # Extract port information
                    if port_col is not None and len(row) > port_col:
                        port = row[port_col]
                        if port is not None:
                            patterns['port_patterns'].add(str(port))
                            port_counts[str(port)] = port_counts.get(str(port), 0) + 1

                    # Extract network features for first 1000 rows (sample)
                    if row_count <= 1000:
                        flow_data = {}
                        for col_idx, col_name in feature_cols[:10]:  # Limit to first 10 features
                            if len(row) > col_idx and row[col_idx] is not None:
                                flow_data[col_name] = row[col_idx]

                        if flow_data:
                            patterns['traffic_patterns'].append({
                                'attack_type': attack_type if attack_type_col is not None else 'Unknown',
                                'features': flow_data
                            })

        elif hasattr(sheet_obj, 'nrows'):  # xlrd
            for row_idx in range(1, min(1001, sheet_obj.nrows)):  # Limit to 1000 rows
                row_count += 1

                # Extract attack type
                if attack_type_col is not None:
                    attack_type = str(sheet_obj.cell_value(row_idx, attack_type_col)).strip()
                    patterns['attack_types'].add(attack_type)
                    attack_counts[attack_type] = attack_counts.get(attack_type, 0) + 1

                # Extract port information
                if port_col is not None:
                    port = sheet_obj.cell_value(row_idx, port_col)
                    if port is not None:
                        patterns['port_patterns'].add(str(port))
                        port_counts[str(port)] = port_counts.get(str(port), 0) + 1

    except Exception as e:
        print(f"⚠️  Error extracting full dataset: {e}")
        print("   Using sample data only...")

        # Fallback to sample data
        for row in data_rows:
            if attack_type_col is not None and len(row) > attack_type_col:
                attack_type = str(row[attack_type_col]).strip()
                patterns['attack_types'].add(attack_type)
                attack_counts[attack_type] = attack_counts.get(attack_type, 0) + 1

            if port_col is not None and len(row) > port_col:
                port = row[port_col]
                if port is not None:
                    patterns['port_patterns'].add(str(port))

    # Convert sets to lists and add statistics
    patterns['attack_types'] = list(patterns['attack_types'])
    patterns['port_patterns'] = list(patterns['port_patterns'])
    patterns['attack_statistics'] = attack_counts
    patterns['port_statistics'] = port_counts

    # Create attack signatures based on the data
    for attack_type in patterns['attack_types']:
        if attack_type.lower() != 'normal traffic':
            patterns['attack_signatures'].append({
                'name': attack_type,
                'description': f"Network traffic pattern for {attack_type}",
                'severity': 'high' if 'dos' in attack_type.lower() or 'attack' in attack_type.lower() else 'medium'
            })

    # Print summary
    print(f"\n📊 Extracted patterns from {row_count} rows:")
    print(f"   • Attack Types: {len(patterns['attack_types'])} unique types")
    if patterns['attack_types']:
        print(f"     Types: {patterns['attack_types']}")

    print(f"   • Port Patterns: {len(patterns['port_patterns'])} unique ports")
    if patterns['port_patterns']:
        top_ports = sorted(port_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        print(f"     Top ports: {[f'{port}({count})' for port, count in top_ports]}")

    print(f"   • Traffic Patterns: {len(patterns['traffic_patterns'])} samples")
    print(f"   • Attack Signatures: {len(patterns['attack_signatures'])} signatures")

    if attack_counts:
        print(f"\n📈 Attack Type Distribution:")
        for attack_type, count in sorted(attack_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"     • {attack_type}: {count} instances")

    return patterns

def save_training_data(patterns):
    """Save extracted patterns to training files"""
    print("\n💾 Saving training data...")

    # Create training directory
    training_dir = "var/lib/cowrie/training_data"
    os.makedirs(training_dir, exist_ok=True)

    # Save patterns as JSON
    patterns_file = os.path.join(training_dir, "network_attack_patterns.json")
    with open(patterns_file, 'w') as f:
        json.dump(patterns, f, indent=2)
    print(f"✅ Updated network attack patterns: {patterns_file}")

    # Save attack types
    if patterns['attack_types']:
        attack_types_file = os.path.join(training_dir, "attack_types.txt")
        with open(attack_types_file, 'w') as f:
            for attack_type in patterns['attack_types']:
                f.write(f"{attack_type}\n")
        print(f"✅ Updated attack types: {attack_types_file}")

    # Save port patterns
    if patterns['port_patterns']:
        ports_file = os.path.join(training_dir, "target_ports.txt")
        with open(ports_file, 'w') as f:
            # Sort ports by frequency if statistics available
            if 'port_statistics' in patterns:
                sorted_ports = sorted(patterns['port_statistics'].items(),
                                    key=lambda x: x[1], reverse=True)
                for port, count in sorted_ports:
                    f.write(f"{port}\t{count}\n")
            else:
                for port in patterns['port_patterns']:
                    f.write(f"{port}\n")
        print(f"✅ Updated target ports: {ports_file}")

    # Save attack signatures
    if patterns['attack_signatures']:
        signatures_file = os.path.join(training_dir, "attack_signatures.json")
        with open(signatures_file, 'w') as f:
            json.dump(patterns['attack_signatures'], f, indent=2)
        print(f"✅ Updated attack signatures: {signatures_file}")

    # Save traffic patterns (sample)
    if patterns['traffic_patterns']:
        traffic_file = os.path.join(training_dir, "traffic_patterns.json")
        with open(traffic_file, 'w') as f:
            # Save first 100 patterns to avoid huge files
            sample_patterns = patterns['traffic_patterns'][:100]
            json.dump(sample_patterns, f, indent=2)
        print(f"✅ Updated traffic patterns: {traffic_file} ({len(sample_patterns)} samples)")

    # Create a summary report
    summary_file = os.path.join(training_dir, "training_summary.txt")
    with open(summary_file, 'w') as f:
        f.write("COWRIE HONEYPOT TRAINING DATA SUMMARY\n")
        f.write("="*50 + "\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

        f.write(f"Attack Types Found: {len(patterns['attack_types'])}\n")
        for attack_type in patterns['attack_types']:
            count = patterns.get('attack_statistics', {}).get(attack_type, 0)
            f.write(f"  • {attack_type}: {count} instances\n")

        f.write(f"\nTarget Ports: {len(patterns['port_patterns'])}\n")
        if 'port_statistics' in patterns:
            top_ports = sorted(patterns['port_statistics'].items(),
                             key=lambda x: x[1], reverse=True)[:10]
            for port, count in top_ports:
                f.write(f"  • Port {port}: {count} connections\n")

        f.write(f"\nTraining Data Files:\n")
        f.write(f"  • network_attack_patterns.json - Complete pattern data\n")
        f.write(f"  • attack_types.txt - List of attack types\n")
        f.write(f"  • target_ports.txt - Targeted ports with frequencies\n")
        f.write(f"  • attack_signatures.json - Attack signature definitions\n")
        f.write(f"  • traffic_patterns.json - Sample network flow patterns\n")

    print(f"✅ Created training summary: {summary_file}")

def main():
    """Main function to read Excel dataset and extract training data"""
    dataset_file = "Cropped dataset.xlsx"
    
    if not os.path.exists(dataset_file):
        print(f"❌ Dataset file not found: {dataset_file}")
        print(f"   Please ensure the file exists in the current directory")
        return 1
    
    print("🎯 EXCEL DATASET READER FOR COWRIE TRAINING")
    print("="*60)
    print(f"📁 Dataset file: {dataset_file}")
    print(f"📊 File size: {os.path.getsize(dataset_file) / (1024*1024):.2f} MB")
    print("="*60)
    
    # Try different methods to read the Excel file
    patterns = None
    
    # Method 1: Try openpyxl (for .xlsx files)
    patterns = try_openpyxl_read(dataset_file)
    
    # Method 2: Try xlrd (for older .xls files)
    if patterns is None:
        patterns = try_xlrd_read(dataset_file)
    
    if patterns is None:
        print("\n❌ Could not read Excel file with available libraries")
        print("   Please install one of the following:")
        print("   • pip install openpyxl (for .xlsx files)")
        print("   • pip install xlrd (for .xls files)")
        return 1
    
    # Save the extracted patterns
    save_training_data(patterns)
    
    print("\n🎉 EXCEL DATASET PROCESSING COMPLETE!")
    print("="*60)
    print("✅ Training data has been extracted and saved")
    print("✅ Cowrie training module will use this real data")
    print("✅ Restart Cowrie to apply the new training patterns")
    
    return 0

if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n\n🛑 Process interrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        sys.exit(1)
